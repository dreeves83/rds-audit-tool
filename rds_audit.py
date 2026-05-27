import boto3
import re
import sys
from datetime import datetime, timezone, timedelta
from botocore.config import Config
from botocore.exceptions import ClientError, ConnectTimeoutError, NoCredentialsError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────
SNAPSHOT_RETENTION_DAYS  = 30
CLOUDWATCH_LOOKBACK_DAYS = 30
BOTO_CONFIG = Config(connect_timeout=10, retries={"max_attempts": 1})
ENVIRONMENTS = ["prod", "qa", "uat"]

MIN_ALLOC_GIB = 50

# Utilization flagging — flag when used% is BELOW these
FLAG_ORANGE_PCT = 25   # conservative threshold — clearest offenders
FLAG_YELLOW_PCT = 50   # moderate threshold

SAVINGS_SCENARIOS = [
    (FLAG_ORANGE_PCT, "Conservative"),
    (FLAG_YELLOW_PCT, "Moderate"),
]

# ── Style ─────────────────────────────────────────────────────────────────────
FONT_HEADER = "Aptos Black"
FONT_DATA   = "Aptos Narrow"
SIZE_HEADER = 12
SIZE_DATA   = 11
FILL_HEADER = "D6E4F7"
FILL_RED    = "FFCCCC"
FILL_ORANGE = "FFD9B3"
FILL_YELLOW = "FFFACD"
FILL_NAVY   = "1F4E79"
FILL_INPUT  = "FFF2CC"
FILL_STALE  = "E8D5F5"  # light purple — suspected stale instance
FONT_WHITE  = "FFFFFF"


# ══════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def no_border():
    return Border()

def make_font(name, size, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def style_cell(cell, fill=None, bold=False, italic=False, align=None):
    cell.font      = make_font(FONT_DATA, SIZE_DATA, bold=bold, italic=italic)
    cell.alignment = align or left()
    cell.border    = thin_border()
    if fill:
        cell.fill = make_fill(fill)

def header_cell(ws, row, col, label):
    c = ws.cell(row=row, column=col, value=label)
    c.font      = make_font(FONT_HEADER, SIZE_HEADER, bold=True)
    c.fill      = make_fill(FILL_HEADER)
    c.alignment = center()
    c.border    = thin_border()
    return c

def title_row(ws, row, label, num_cols):
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=1, value=label)
    c.font      = make_font(FONT_HEADER, 14, bold=True, color=FONT_WHITE)
    c.fill      = make_fill(FILL_NAVY)
    c.alignment = left()
    c.border    = thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)

def section_label(ws, row, label, num_cols):
    c = ws.cell(row=row, column=1, value=label)
    c.font      = make_font(FONT_HEADER, SIZE_HEADER, bold=True, color=FONT_WHITE)
    c.fill      = make_fill(FILL_NAVY)
    c.alignment = left()
    c.border    = thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    ws.row_dimensions[row].height = 20
    return row + 1

def bool_fmt(val):
    if val is True:  return "Yes"
    if val is False: return "No"
    return "—"

def fmt_date(iso_str):
    try:
        dt = datetime.fromisoformat(str(iso_str).replace("Z", "+00:00"))
        day    = dt.day
        suffix = "th" if 11 <= day <= 13 else {1:"st",2:"nd",3:"rd"}.get(day % 10, "th")
        return dt.strftime(f"%B {day}{suffix} %Y")
    except Exception:
        return str(iso_str)

def is_stale(identifier):
    """
    Flag any instance that doesn't match the standard env-XX-database pattern.
    Read replicas are excluded as they are legitimate infrastructure.
    """
    import re
    iid = identifier.lower()
    if 'readreplica' in iid:
        return False
    return not bool(re.match(r'^(prod|qa|uat)-\d+-database$', iid))


def util_fill(allocated, used):
    if not allocated or allocated < MIN_ALLOC_GIB or used is None:
        return None
    pct = (used / allocated) * 100
    if pct < FLAG_ORANGE_PCT:
        return FILL_ORANGE
    if pct < FLAG_YELLOW_PCT:
        return FILL_YELLOW
    return None

def get_util_fill(allocated, used_pct):
    if not allocated or allocated < MIN_ALLOC_GIB or used_pct is None:
        return None
    if used_pct < FLAG_ORANGE_PCT:
        return FILL_ORANGE
    if used_pct < FLAG_YELLOW_PCT:
        return FILL_YELLOW
    return None

def compute_summary(env_data):
    """Derive all summary values from data at runtime. No hardcoding."""
    total_instances    = 0
    total_allocated    = 0.0
    total_used         = 0.0
    no_auto_backup     = 0
    suspected_stale    = 0
    orphaned_snapshots = 0
    safe_to_delete     = 0
    excess_con         = 0.0
    excess_mod         = 0.0

    for env, edata in env_data.items():
        for rdata in edata["Regions"]:
            for inst in rdata["Instances"]:
                total_instances += 1
                alloc = inst["AllocatedStorageGiB"] or 0
                used  = inst["UsedStorageGiB"]  or 0
                pct   = inst["UsedStoragePct"]   or 0
                total_allocated += alloc
                total_used      += used

                if not inst["AutoBackupEnabled"]:
                    no_auto_backup += 1

                if is_stale(inst["DBInstanceIdentifier"]):
                    suspected_stale += 1

                if alloc >= MIN_ALLOC_GIB:
                    if pct < FLAG_ORANGE_PCT:
                        excess_con += max(0, alloc - (used / (FLAG_ORANGE_PCT / 100)))
                        excess_mod += max(0, alloc - (used / (FLAG_YELLOW_PCT / 100)))
                    elif pct < FLAG_YELLOW_PCT:
                        excess_mod += max(0, alloc - (used / (FLAG_YELLOW_PCT / 100)))

                for snap in inst.get("ManualSnapshots", []):
                    if snap["Status"] == "Safe to Delete":
                        safe_to_delete += 1

            orphaned_snapshots += len(rdata.get("Orphans", []))

    return {
        "TotalInstances":        total_instances,
        "TotalAllocatedGiB":     round(total_allocated, 2),
        "TotalUsedGiB":          round(total_used, 2),
        "NoAutoBackup":          no_auto_backup,
        "SuspectedStale":        suspected_stale,
        "OrphanedSnapshots":     orphaned_snapshots,
        "SafeToDelete":          safe_to_delete,
        "ExcessGiBConservative": round(excess_con, 2),
        "ExcessGiBModerate":     round(excess_mod, 2),
    }

def fit_columns(ws, col_headers, data_rows, start_col=1):
    """
    Set column widths. Header text length is always the minimum floor.
    Data values can only expand, never shrink below header width.
    """
    widths = {}
    for i, h in enumerate(col_headers):
        col = start_col + i
        widths[col] = len(str(h)) + 6  # header is the floor with generous padding

    for row_vals in data_rows:
        for i, val in enumerate(row_vals):
            col = start_col + i
            if val is not None:
                widths[col] = max(widths[col], len(str(val)) + 4)

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width, 80)


# ══════════════════════════════════════════════════════════════════════════════
# AWS HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def paste_aws_exports():
    print("\nPaste the AWS export block below.")
    print("Press ENTER on a blank line when done.\n")
    lines = []
    while True:
        line = input()
        if line.strip() == "":
            break
        lines.append(line)
    text = "\n".join(lines)
    creds = {}
    for key in ("AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY", "AWS_SESSION_TOKEN"):
        m = re.search(rf'{key}\s*=\s*["\']?([^"\']+)["\']?', text)
        if m:
            creds[key] = m.group(1).strip()
    missing = [k for k in ("AWS_ACCESS_KEY_ID","AWS_SECRET_ACCESS_KEY","AWS_SESSION_TOKEN")
               if k not in creds]
    if missing:
        print("\nMissing credentials:")
        for k in missing: print(f"  - {k}")
        sys.exit(1)
    return creds

def validate_credentials(session):
    try:
        identity = session.client("sts").get_caller_identity()
        print(f"  Authenticated — Account: {identity.get('Account')}")
        return True
    except (ClientError, NoCredentialsError) as e:
        print(f"  Authentication failed: {e}")
        return False

def get_enabled_regions(session):
    ec2 = session.client("ec2", region_name="us-east-1")
    try:
        r = ec2.describe_regions(Filters=[{"Name":"opt-in-status",
            "Values":["opt-in-not-required","opted-in"]}])
        return sorted([x["RegionName"] for x in r["Regions"]])
    except ClientError as e:
        print(f"Failed to retrieve regions: {e}"); sys.exit(1)

def get_rds_instances(session, region):
    rds = session.client("rds", region_name=region, config=BOTO_CONFIG)
    instances = []
    try:
        for page in rds.get_paginator("describe_db_instances").paginate():
            instances.extend(page["DBInstances"])
    except (ClientError, ConnectTimeoutError) as e:
        print(f"  Warning: {region} — {type(e).__name__}")
    return instances

def get_all_snapshots(session, region):
    rds = session.client("rds", region_name=region, config=BOTO_CONFIG)
    snapshots = []
    try:
        for page in rds.get_paginator("describe_db_snapshots").paginate():
            snapshots.extend(page["DBSnapshots"])
    except (ClientError, ConnectTimeoutError) as e:
        print(f"  Warning: {region} snapshots — {type(e).__name__}")
    return snapshots

def get_free_storage_avg(cw, identifier):
    end   = datetime.now(timezone.utc)
    start = end - timedelta(days=CLOUDWATCH_LOOKBACK_DAYS)
    try:
        r = cw.get_metric_statistics(
            Namespace="AWS/RDS", MetricName="FreeStorageSpace",
            Dimensions=[{"Name":"DBInstanceIdentifier","Value":identifier}],
            StartTime=start, EndTime=end,
            Period=int(timedelta(days=CLOUDWATCH_LOOKBACK_DAYS).total_seconds()),
            Statistics=["Average"],
        )
        pts = r.get("Datapoints", [])
        return round(pts[0]["Average"] / (1024**3), 2) if pts else None
    except ClientError:
        return None

def stack_sort_key(identifier):
    m = re.search(r'[a-z]+-(\d+)', identifier)
    return int(m.group(1)) if m else 999999

def extract_env_prefix(identifier):
    m = re.match(r'([a-z]+-\d+)-', identifier)
    return m.group(0) if m else None

def parse_snap_time(raw):
    if isinstance(raw, datetime):
        dt = raw
    else:
        dt = datetime.fromisoformat(str(raw))
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


# ══════════════════════════════════════════════════════════════════════════════
# AUDIT LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def audit_region(session, region):
    instances = get_rds_instances(session, region)
    if not instances:
        return None

    print(f"    {region} — {len(instances)} instances")
    snapshots    = get_all_snapshots(session, region)
    manual_snaps = [s for s in snapshots if s["SnapshotType"] == "manual"]
    auto_count   = len(snapshots) - len(manual_snaps)

    cw  = session.client("cloudwatch", region_name=region, config=BOTO_CONFIG)
    now = datetime.now(timezone.utc)

    prefix_map = {}
    for inst in instances:
        p = extract_env_prefix(inst["DBInstanceIdentifier"])
        if p and p not in prefix_map:
            prefix_map[p] = inst

    snap_map = {}
    orphans  = []

    for snap in manual_snaps:
        snap_id  = snap["DBSnapshotIdentifier"]
        created  = parse_snap_time(snap["SnapshotCreateTime"])
        age_days = (now - created).days
        exceeds  = age_days > SNAPSHOT_RETENTION_DAYS
        prefix   = extract_env_prefix(snap_id)
        matched  = prefix_map.get(prefix) if prefix else None

        rec = {
            "SnapshotIdentifier": snap_id,
            "CreatedDate":        created.strftime("%b %d %Y"),
            "AgeDays":            age_days,
            "ExceedsRetention":   exceeds,
        }

        if not matched:
            rec["Status"] = "Needs Review"
            orphans.append(rec)
        else:
            mid     = matched["DBInstanceIdentifier"]
            auto_bk = matched.get("BackupRetentionPeriod", 0) > 0
            if exceeds and auto_bk:       status = "Safe to Delete"
            elif exceeds and not auto_bk: status = "Needs Review"
            else:                         status = "Within Retention"
            rec["Status"]          = status
            rec["MatchedInstance"] = mid
            snap_map.setdefault(mid, []).append(rec)

    records = []
    for inst in sorted(instances, key=lambda x: stack_sort_key(x["DBInstanceIdentifier"])):
        iid         = inst["DBInstanceIdentifier"]
        allocated   = inst.get("AllocatedStorage", 0)
        auto_backup = inst.get("BackupRetentionPeriod", 0) > 0

        free_gib = get_free_storage_avg(cw, iid)
        if free_gib is not None:
            used_gib = round(allocated - free_gib, 2)
            used_pct = round((used_gib / allocated) * 100, 1) if allocated else None
            excess   = round(free_gib, 2)
        else:
            used_gib = used_pct = excess = None

        inst_snaps   = snap_map.get(iid, [])
        safe_count   = sum(1 for s in inst_snaps if s["Status"] == "Safe to Delete")
        review_count = sum(1 for s in inst_snaps if s["Status"] == "Needs Review")

        records.append({
            "DBInstanceIdentifier": iid,
            "AutoBackupEnabled":    auto_backup,
            "AllocatedStorageGiB":  allocated,
            "UsedStorageGiB":       used_gib,
            "FreeStorageGiB":       free_gib,
            "UsedStoragePct":       used_pct,
            "ExcessStorageGiB":     excess,
            "ManualSnapshots":      inst_snaps,
            "SafeToDeleteCount":    safe_count,
            "NeedsReviewCount":     review_count,
        })

    def excess_gib_for_threshold(threshold_pct):
        total = 0
        for r in records:
            alloc = r["AllocatedStorageGiB"]
            used  = r["UsedStorageGiB"] or 0
            if alloc < MIN_ALLOC_GIB or r["UsedStoragePct"] is None:
                continue
            if r["UsedStoragePct"] < threshold_pct:
                right_sized = used / (threshold_pct / 100)
                total += max(0, alloc - right_sized)
        return round(total, 2)

    return {
        "Region":   region,
        "Summary": {
            "TotalInstances":        len(records),
            "TotalAllocatedGiB":     round(sum(r["AllocatedStorageGiB"] for r in records), 2),
            "TotalUsedGiB":          round(sum(r["UsedStorageGiB"] or 0 for r in records), 2),
            "ExcessGiBConservative": excess_gib_for_threshold(FLAG_ORANGE_PCT),
            "ExcessGiBModerate":     excess_gib_for_threshold(FLAG_YELLOW_PCT),
            "ManualSnapshots":       len(manual_snaps),
            "AutomatedSnapshots":    auto_count,
            "SafeToDelete":          sum(r["SafeToDeleteCount"] for r in records),
            "NeedsReview":           sum(r["NeedsReviewCount"] for r in records),
            "OrphanedSnapshots":     len(orphans),
            "NoAutoBackup":          sum(1 for r in records if not r["AutoBackupEnabled"]),
            "SuspectedStale":        sum(1 for r in records if is_stale(r["DBInstanceIdentifier"])),
        },
        "Instances": records,
        "Snapshots": [s for r in records for s in r["ManualSnapshots"]],
        "Orphans":   orphans,
    }


def audit_environment(env_name, creds):
    print(f"\n  {env_name.upper()}")
    session = boto3.Session(
        aws_access_key_id=creds["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=creds["AWS_SECRET_ACCESS_KEY"],
        aws_session_token=creds["AWS_SESSION_TOKEN"],
        region_name="us-east-1",
    )
    if not validate_credentials(session):
        return None
    all_regions    = get_enabled_regions(session)
    region_results = []
    for region in all_regions:
        try:
            result = audit_region(session, region)
            if result:
                region_results.append(result)
        except (ConnectTimeoutError, Exception) as e:
            print(f"    Skipping {region}: {type(e).__name__}")
    return region_results if region_results else None


def build_global_summary(regions):
    g = {k: 0 for k in ["TotalInstances","TotalAllocatedGiB","TotalUsedGiB",
                          "ExcessGiBConservative","ExcessGiBModerate",
                          "ManualSnapshots","SafeToDelete","NeedsReview",
                          "OrphanedSnapshots","NoAutoBackup","SuspectedStale"]}
    g["TotalRegions"] = len(regions)
    for r in regions:
        s = r["Summary"]
        for k in g:
            if k != "TotalRegions":
                g[k] = round(g[k] + s.get(k, 0), 2)
    return g


def calc_savings(excess_gib, rate=0.115):
    return round(excess_gib * rate * 12)


# ══════════════════════════════════════════════════════════════════════════════
# XLSX — INSTANCE TAB
# ══════════════════════════════════════════════════════════════════════════════

INST_HEADERS = [
    "RDS Instance", "Automated Backup", "Allocated (GiB)",
    "Used (GiB)", "Free (GiB)", "Used %", "Excess (GiB)"
]

SNAP_HEADERS = [
    "RDS Instance", "Snapshot", "Created",
    "Age (Days)", "Exceeds 30 Days", "Status"
]

ORPHAN_HEADERS = ["Snapshot", "Created", "Age (Days)", "Status"]


def build_instance_sheet(wb, env, region_data):
    region    = region_data["Region"]
    instances = region_data["Instances"]
    NC        = len(INST_HEADERS)

    ws = wb.create_sheet(title=f"{env.upper()} {region}"[:31])
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    title_row(ws, 1, f"RDS Instances - {env.upper()} / {region}", NC)
    for ci, label in enumerate(INST_HEADERS, start=1):
        header_cell(ws, 2, ci, label)

    data_rows = []
    for row_idx, inst in enumerate(instances, start=3):
        ws.row_dimensions[row_idx].height = 18
        no_backup  = not inst["AutoBackupEnabled"]
        stale      = is_stale(inst["DBInstanceIdentifier"])
        uf         = get_util_fill(inst["AllocatedStorageGiB"], inst["UsedStoragePct"])

        vals = [
            inst["DBInstanceIdentifier"],
            bool_fmt(inst["AutoBackupEnabled"]),
            inst["AllocatedStorageGiB"],
            inst["UsedStorageGiB"],
            inst["FreeStorageGiB"],
            inst["UsedStoragePct"],
            inst["ExcessStorageGiB"],
        ]
        fills = [
            FILL_RED if no_backup else (FILL_STALE if stale else None),  # RDS Instance name
            FILL_RED if no_backup else None,  # Automated Backup
            None,                             # Allocated (GiB)
            None,                             # Used (GiB)
            None,                             # Free (GiB)
            uf,                               # Used % — same as Excess
            uf,                               # Excess (GiB)
        ]
        for ci, (val, fill) in enumerate(zip(vals, fills), start=1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            style_cell(c, fill=fill, bold=(ci == 1))
        data_rows.append(vals)

    fit_columns(ws, INST_HEADERS, data_rows)

    # Legend — placed in columns NC+2 and NC+3, narrow swatch + description
    lc = NC + 2
    legend_items = [
        (None,        "Legend"),
        (FILL_RED,    f"Automated backup disabled"),
        (FILL_ORANGE, f"{FLAG_ORANGE_PCT}% or less of allocated storage in use - strong right-sizing candidate"),
        (FILL_YELLOW, f"Between {FLAG_ORANGE_PCT}% and {FLAG_YELLOW_PCT}% of allocated storage in use - monitor"),
        (FILL_STALE,  "Suspected stale instance (old, temp, restored, or flagged for review)"),
    ]
    max_desc = max(len(item[1]) for item in legend_items)
    for i, (fill, desc) in enumerate(legend_items):
        r = 2 + i
        ws.row_dimensions[r].height = 18
        c = ws.cell(row=r, column=lc, value=desc)
        c.border    = thin_border()
        c.alignment = left()
        c.font      = make_font(FONT_DATA if fill else FONT_HEADER,
                                SIZE_DATA if fill else SIZE_HEADER,
                                bold=(fill is None))
        c.fill      = make_fill(fill if fill else FILL_HEADER)

    ws.column_dimensions[get_column_letter(lc)].width = max_desc + 4


def build_snapshot_sheet(wb, env, region_data):
    region    = region_data["Region"]
    snapshots = region_data.get("Snapshots", [])
    if not snapshots:
        return

    NC = len(SNAP_HEADERS)
    ws = wb.create_sheet(title=f"{env.upper()} {region} Snapshots"[:31])
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    title_row(ws, 1, f"Manual Snapshots - {env.upper()} / {region}", NC)
    for ci, label in enumerate(SNAP_HEADERS, start=1):
        header_cell(ws, 2, ci, label)

    data_rows = []
    for row_idx, snap in enumerate(snapshots, start=3):
        ws.row_dimensions[row_idx].height = 18
        status = snap["Status"]
        fill   = FILL_RED    if status == "Needs Review"   else (
                 FILL_YELLOW if status == "Safe to Delete"  else None)
        vals = [
            snap.get("MatchedInstance", "-"),
            snap["SnapshotIdentifier"],
            snap["CreatedDate"],
            snap["AgeDays"],
            bool_fmt(snap["ExceedsRetention"]),
            status,
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            style_cell(c, fill=fill)
        data_rows.append(vals)

    fit_columns(ws, SNAP_HEADERS, data_rows)

    lc = NC + 2
    legend_items = [
        (None,        "Legend"),
        (FILL_YELLOW, "Safe to Delete - exceeds 30 days, automated backup confirmed"),
        (FILL_RED,    "Needs Review - exceeds 30 days, automated backup not confirmed"),
    ]
    max_desc = max(len(item[1]) for item in legend_items)
    for i, (fill, desc) in enumerate(legend_items):
        r = 2 + i
        ws.row_dimensions[r].height = 18
        c = ws.cell(row=r, column=lc, value=desc)
        c.border = thin_border(); c.alignment = left()
        c.font   = make_font(FONT_DATA if fill else FONT_HEADER,
                             SIZE_DATA if fill else SIZE_HEADER,
                             bold=(fill is None))
        c.fill   = make_fill(fill if fill else FILL_HEADER)

    ws.column_dimensions[get_column_letter(lc)].width = max_desc + 4


def build_orphan_sheet(wb, env, region_data):
    region  = region_data["Region"]
    orphans = region_data.get("Orphans", [])
    if not orphans:
        return

    NC = len(ORPHAN_HEADERS)
    ws = wb.create_sheet(title=f"{env.upper()} {region} Orphans"[:31])
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    title_row(ws, 1, f"Orphaned Snapshots - {env.upper()} / {region}", NC)
    for ci, label in enumerate(ORPHAN_HEADERS, start=1):
        header_cell(ws, 2, ci, label)

    data_rows = []
    for row_idx, snap in enumerate(orphans, start=3):
        ws.row_dimensions[row_idx].height = 18
        vals = [
            snap["SnapshotIdentifier"],
            snap["CreatedDate"],
            snap["AgeDays"],
            snap["Status"],
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            style_cell(c, fill=FILL_RED)
        data_rows.append(vals)

    fit_columns(ws, ORPHAN_HEADERS, data_rows)

    lc = NC + 2
    legend_items = [
        (None,     "Legend"),
        (FILL_RED, "No living RDS instance found matching this snapshot"),
    ]
    max_desc = max(len(item[1]) for item in legend_items)
    for i, (fill, desc) in enumerate(legend_items):
        r = 2 + i
        ws.row_dimensions[r].height = 18
        c = ws.cell(row=r, column=lc, value=desc)
        c.border = thin_border(); c.alignment = left()
        c.font   = make_font(FONT_DATA if fill else FONT_HEADER,
                             SIZE_DATA if fill else SIZE_HEADER,
                             bold=(fill is None))
        c.fill   = make_fill(fill if fill else FILL_HEADER)

    ws.column_dimensions[get_column_letter(lc)].width = max_desc + 4


# ══════════════════════════════════════════════════════════════════════════════
# XLSX — SUMMARY TAB
# ══════════════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, audit_ts, summary):
    ws        = wb.create_sheet(title="Summary", index=0)
    DATA_COLS = 2
    row       = 1

    title_row(ws, row, "RDS Multi-Region Audit Report - Prod, QA, and UAT", DATA_COLS)
    row += 1
    c = ws.cell(row=row, column=1, value=f"Generated: {fmt_date(audit_ts)}")
    c.font = make_font(FONT_DATA, SIZE_DATA, italic=True); c.alignment = left()
    ws.row_dimensions[row].height = 18
    row += 2

    exc_con_row = None
    exc_mod_row = None

    def kv(label, value, fill=None):
        nonlocal row
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = make_font(FONT_DATA, SIZE_DATA, bold=True)
        lc.alignment = left(); lc.border = thin_border()
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = make_font(FONT_DATA, SIZE_DATA)
        vc.alignment = center(); vc.border = thin_border()
        if fill:
            lc.fill = make_fill(fill)
            vc.fill = make_fill(fill)
        r = row
        row += 1
        return r

    # Snapshot Findings
    row = section_label(ws, row, "Snapshot Findings", DATA_COLS)
    kv("Orphaned Snapshots (no living RDS instance found)",
       summary["OrphanedSnapshots"],
       fill=FILL_RED if summary["OrphanedSnapshots"] > 0 else None)
    kv("Manual Snapshots Safe to Delete (exceeds 30 days, automated backup confirmed)",
       summary["SafeToDelete"],
       fill=FILL_YELLOW if summary["SafeToDelete"] > 0 else None)
    row += 1

    # RDS Instance Findings
    row = section_label(ws, row, "RDS Instance Findings", DATA_COLS)
    kv("Total Running RDS Instances", summary["TotalInstances"])
    kv("Instances Without Automated Backup",
       summary["NoAutoBackup"],
       fill=FILL_RED if summary["NoAutoBackup"] > 0 else None)
    kv("Suspected Stale Instances (old, temp, restored, or flagged for review)",
       summary["SuspectedStale"],
       fill=FILL_STALE if summary["SuspectedStale"] > 0 else None)
    row += 1

    # Storage Allocation
    row = section_label(ws, row, "RDS Instance Storage Allocation", DATA_COLS)
    kv("Total Allocated (GiB)", summary["TotalAllocatedGiB"])
    kv("Total Used (GiB)",      summary["TotalUsedGiB"])
    exc_con_row = kv(
        f"Excess - Conservative (instances using {FLAG_ORANGE_PCT}% or less of allocated)",
        summary["ExcessGiBConservative"])
    exc_mod_row = kv(
        f"Excess - Moderate (instances using {FLAG_YELLOW_PCT}% or less of allocated)",
        summary["ExcessGiBModerate"])
    row += 1

    # Right-Sizing Savings Estimate
    row = section_label(ws, row, "Right-Sizing Savings Estimate", DATA_COLS)

    note = ws.cell(row=row, column=1,
        value="Verify current gp3 RDS storage rate at https://aws.amazon.com/rds/pricing/")
    note.font = make_font(FONT_DATA, SIZE_DATA, italic=True, color="555555")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DATA_COLS)
    ws.row_dimensions[row].height = 36
    row += 1

    ws.row_dimensions[row].height = 18
    lc = ws.cell(row=row, column=1, value="gp3 Storage Rate ($/GiB-Month):")
    lc.font = make_font(FONT_DATA, SIZE_DATA, bold=True); lc.alignment = left()
    lc.border = thin_border()
    ic = ws.cell(row=row, column=2, value=0.115)
    ic.font = make_font(FONT_DATA, SIZE_DATA, bold=True)
    ic.fill = make_fill(FILL_INPUT); ic.alignment = center()
    ic.number_format = "$#,##0.000"; ic.border = thin_border()
    rate_row = row
    row += 1

    header_cell(ws, row, 1, "Scenario")
    header_cell(ws, row, 2, "Est. Annual Savings")
    ws.row_dimensions[row].height = 18
    row += 1

    for excess_row, (threshold_pct, label) in zip(
            [exc_con_row, exc_mod_row], SAVINGS_SCENARIOS):
        ws.row_dimensions[row].height = 18
        desc = f"{label} - instances using {threshold_pct}% or less of allocated storage"
        lc2 = ws.cell(row=row, column=1, value=desc)
        lc2.font = make_font(FONT_DATA, SIZE_DATA); lc2.alignment = left()
        lc2.border = thin_border()
        vc2 = ws.cell(row=row, column=2, value=f"=B{excess_row}*B{rate_row}*12")
        vc2.font = make_font(FONT_DATA, SIZE_DATA, bold=True)
        vc2.alignment = center(); vc2.border = thin_border()
        vc2.number_format = "$#,##0"
        row += 1

    all_labels = [
        "Orphaned Snapshots (no living RDS instance found)",
        "Manual Snapshots Safe to Delete (exceeds 30 days, automated backup confirmed)",
        "Total Running RDS Instances",
        "Instances Without Automated Backup",
        "Total Allocated (GiB)",
        "Total Used (GiB)",
        f"Excess - Conservative (instances using {FLAG_ORANGE_PCT}% or less of allocated)",
        f"Excess - Moderate (instances using {FLAG_YELLOW_PCT}% or less of allocated)",
        "gp3 Storage Rate ($/GiB-Month):",
        f"Conservative - instances using {FLAG_ORANGE_PCT}% or less of allocated storage",
        f"Moderate - instances using {FLAG_YELLOW_PCT}% or less of allocated storage",
        "Verify current gp3 RDS storage rate at https://aws.amazon.com/rds/pricing/",
    ]
    ws.column_dimensions["A"].width = max(len(l) for l in all_labels) + 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3


# ══════════════════════════════════════════════════════════════════════════════
# XLSX BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_xlsx(data, output_file):
    audit_ts = data["AuditTimestamp"]
    env_data = data["Environments"]

    summary = compute_summary(env_data)

    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, audit_ts, summary)

    for env, edata in env_data.items():
        for rdata in edata["Regions"]:
            if "Snapshots" not in rdata:
                rdata["Snapshots"] = [s for inst in rdata["Instances"]
                                       for s in inst.get("ManualSnapshots", [])]
            build_instance_sheet(wb, env, rdata)
            build_snapshot_sheet(wb, env, rdata)
            build_orphan_sheet(wb, env, rdata)
            print(f"  {env.upper()} / {rdata['Region']} - "
                  f"{len(rdata['Instances'])} instances, "
                  f"{len(rdata.get('Orphans', []))} orphans")

    wb.save(output_file)
    print(f"\n  Report saved to: {output_file}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("\nRDS Audit — Storage, Snapshots & Excess")
    print("----------------------------------------")
    print("Read-only. Credentials used for this run only, never saved.")
    print(f"\nThis audit covers: {', '.join(e.upper() for e in ENVIRONMENTS)}")

    output_file = input("\nOutput filename [leave blank to auto-generate]: ").strip()
    if not output_file:
        ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"rds_audit_{ts}"
    output_file = re.sub(r'\.(xlsx|json)$', '', output_file)

    all_creds = {}
    for env in ENVIRONMENTS:
        print(f"\n{'─'*50}")
        print(f"  Credentials for {env.upper()}")
        print(f"{'─'*50}")
        all_creds[env] = paste_aws_exports()

    print(f"\nAll credentials collected. Starting audit...")

    all_env_results = {}
    for env in ENVIRONMENTS:
        result = audit_environment(env, all_creds[env])
        if result:
            all_env_results[env] = result

    if not all_env_results:
        print("\nNo RDS instances found.")
        sys.exit(0)

    env_summaries    = {env: build_global_summary(regions)
                        for env, regions in all_env_results.items()}
    all_regions_flat = [r for regions in all_env_results.values() for r in regions]
    global_summary   = build_global_summary(all_regions_flat)
    audit_ts         = datetime.now(timezone.utc).isoformat()

    env_data = {env: {"Summary": env_summaries[env], "Regions": regions}
                for env, regions in all_env_results.items()}

    output = {
        "AuditTimestamp": audit_ts,
        "GlobalSummary":  global_summary,
        "Environments":   env_data,
    }

    import json
    json_file = output_file + ".json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=4, default=str)
    print(f"\n  JSON saved to: {json_file}")

    print(f"\nBuilding report...")
    build_xlsx(output, output_file + ".xlsx")

    g = global_summary
    print(f"\n{'='*50}")
    print(f"  AUDIT COMPLETE")
    print(f"  Running Instances:   {g['TotalInstances']}")
    print(f"  Allocated:           {g['TotalAllocatedGiB']} GiB")
    print(f"  Used:                {g['TotalUsedGiB']} GiB")
    print(f"  Excess Conservative: {g['ExcessGiBConservative']} GiB")
    print(f"  Excess Moderate:     {g['ExcessGiBModerate']} GiB")
    print(f"  Safe to Delete:      {g['SafeToDelete']}")
    print(f"  Orphaned:            {g['OrphanedSnapshots']}")
    print(f"  No Auto Backup:      {g['NoAutoBackup']}")


if __name__ == "__main__":
    main()
